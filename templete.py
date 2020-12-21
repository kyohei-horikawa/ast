import openpyxl
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Border, Side


def set_word(row, column, word):
    ws.cell(row, column).value = word
    ws.cell(row, column).alignment = alignment


def merge(s_r, s_c, e_r, e_c):
    ws.merge_cells(start_row=s_r, start_column=s_c, end_row=e_r, end_column=e_c)


wb = openpyxl.load_workbook("Book1.xlsx")
ws = wb.worksheets[0]

alpha = [chr(ord("a") + i) for i in range(26)]


# alignment,border設定
alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
side = Side(style='medium', color='FF000000')
border = Border(top=side, bottom=side, left=side, right=side)

month = 10
day = 16

for row_cnt in range(31):

    # 日付の設定
    merge(1, row_cnt * 10 + 2, 1, row_cnt * 10 + 11)
    ws.cell(row=1, column=row_cnt * 10 + 2).value = f'{month}月{day}日木曜日'
    ws.cell(row=1, column=row_cnt * 10 + 2).alignment = alignment

    day += 1

    if day > 31:
        day = 1
        month += 1

    # 講師名欄などをmerge
    merge(2, row_cnt * 10 + 2, 3, row_cnt * 10 + 3)
    merge(2, row_cnt * 10 + 4, 2, row_cnt * 10 + 7)
    merge(2, row_cnt * 10 + 8, 2, row_cnt * 10 + 11)
    merge(3, row_cnt * 10 + 6, 3, row_cnt * 10 + 7)
    merge(3, row_cnt * 10 + 10, 3, row_cnt * 10 + 11)

    border = Border(top=side, bottom=side, left=side, right=side)
    for i in range(11):
        ws.cell(row=1, column=row_cnt * 10 + i + 1).border = border

    for col_cnt in range(6):  # 縦のloop

        # cellのmerge

        for i in range(4, 24):
            ws.merge_cells(start_row=col_cnt * 21 + i, start_column=row_cnt * 10 + 2, end_row=col_cnt * 21 + i, end_column=row_cnt * 10 + 3)
            ws.merge_cells(start_row=col_cnt * 21 + i, start_column=row_cnt * 10 + 6, end_row=col_cnt * 21 + i, end_column=row_cnt * 10 + 7)
            ws.merge_cells(start_row=col_cnt * 21 + i, start_column=row_cnt * 10 + 10, end_row=col_cnt * 21 + i, end_column=row_cnt * 10 + 11)

        ws.merge_cells(start_row=col_cnt * 21 + 4, start_column=row_cnt * 10 + 1, end_row=col_cnt * 21 + 24, end_column=row_cnt * 10 + 1)
        ws.merge_cells(start_row=col_cnt * 21 + 24, start_column=row_cnt * 10 + 2, end_row=col_cnt * 21 + 24, end_column=row_cnt * 10 + 11)

        # 文字の挿入

        input = [['10時30分〜12時00分   (1)', '12時50分〜14時20分   (2)', '14時40分〜16時10分   (3)', '16時30分〜18時00分   (A)', '18時20分〜19時50分   (B)', '20時10分〜21時40分   (C)'], '講師名', '生徒1', '生徒2', '学年', '科目', '生徒名', '学年', '科目', '生徒名']
        locations = ['A4', 'B2', 'D2', 'H2', 'D3', 'E3', 'F3', 'H3', 'I3', 'J3']

        # 授業時間の設定
        if row_cnt == 0 and col_cnt == 0:
            ws[f'a{col_cnt * 21 + 4}'].value = input[0][col_cnt]
            ws[f'a{col_cnt * 21 + 4}'].alignment = alignment
            merge(2, row_cnt * 10 + 1, 3, row_cnt * 10 + 1)

        # 講師名
        set_word(2, row_cnt * 10 + 2, '講師名')

        # 生徒1
        set_word(2, row_cnt * 10 + 4, '生徒1')

        # 生徒2
        set_word(2, row_cnt * 10 + 8, '生徒2')

        # 学年
        set_word(3, row_cnt * 10 + 4, '学年')

        # 科目
        set_word(3, row_cnt * 10 + 5, '科目')

        # 生徒名
        set_word(3, row_cnt * 10 + 6, '生徒名')

        # 学年
        set_word(3, row_cnt * 10 + 8, '学年')

        # 科目
        set_word(3, row_cnt * 10 + 9, '科目')

        # 生徒名
        set_word(3, row_cnt * 10 + 10, '生徒名')

        # widthの設定
        ws.row_dimensions[col_cnt * 21 + 24].height = 60

        if col_cnt == 0:
            ws.column_dimensions['A'].width = 3.7

            for i, word in enumerate(alpha):
                if 0 < i and i < 11:
                    ws.column_dimensions[word].width = 6

        # 囲みの設定
        side = Side(style='medium', color='FF000000')
        border = Border(top=side, bottom=side, left=side, right=side)

        # 時間を囲む
        for i in range(2, 25):
            ws[f'A{col_cnt * 21 + i}'].border = border

        # 講師名などを囲む
        if col_cnt == 0:
            for i, word in enumerate(alpha):
                if 0 < i and i < 11:
                    for j in range(2, 4):
                        ws[f'{word}{j}'].border = border

        for i, word in enumerate(alpha):
            if i < 11:
                ws[f'{word}{col_cnt * 21 + 24}'].border = border

        side2 = Side(style='thin', color='FF000000')
        border = Border(top=side2, bottom=side2, left=side, right=side)
        for i, word in enumerate(alpha):
            if i < 11:
                for j in range(4, 24):
                    ws[f'{word}{col_cnt * 21 + j}'].border = border


# 保存
wb.save("Book1.xlsx")
